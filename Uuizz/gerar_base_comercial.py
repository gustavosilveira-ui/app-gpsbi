from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "Basegerencial.xlsx"
if not SOURCE.exists():
    SOURCE = ROOT / "pivot.xlsx"
TARGET = ROOT / "base_comercial_misterwiz.js"


def serializable(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return value


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook["Comercial"] if "Comercial" in workbook.sheetnames else workbook.active
header_row = None
headers = None
for row_number, values in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
    if "Data de Inclusão (completa)" in values:
        header_row = row_number
        headers = values
        break
if not header_row or not headers:
    raise RuntimeError("Cabeçalho 'Data de Inclusão (completa)' não encontrado nas 10 primeiras linhas.")
rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)

data = []
for values in rows:
    record = {
        header: serializable(value)
        for header, value in zip(headers, values)
        if header and value is not None
    }
    if record.get("Data de Inclusão (completa)"):
        data.append(record)

payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
TARGET.write_text(f"window.MISTER_WIZ_COMMERCIAL_DATA={payload};\n", encoding="utf-8")
print(f"{len(data)} linhas gravadas em {TARGET.name}")
