from __future__ import annotations
import json
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / 'BaseDados.xlsx'
TARGET = ROOT / 'base_comercial_grupo.js'
ALLOWED_GROUPS = {'EA', 'CW'}


def serializable(value):
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    return value

if not SOURCE.exists():
    raise FileNotFoundError(f'Arquivo não encontrado: {SOURCE.name}')

wb = load_workbook(SOURCE, read_only=True, data_only=True)
if 'BaseGeral' not in wb.sheetnames:
    raise RuntimeError("Aba 'BaseGeral' não encontrada na BaseDados.xlsx")
ws = wb['BaseGeral']
rows = ws.iter_rows(values_only=True)
headers = next(rows)

data = []
for values in rows:
    record = {h: serializable(v) for h, v in zip(headers, values) if h and v is not None}
    group = str(record.get('Grupo') or '').strip().upper()
    if group not in ALLOWED_GROUPS:
        continue
    if not record.get('Data de venda'):
        continue
    data.append(record)

payload = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
TARGET.write_text(f'window.UUIZZ_GROUP_COMMERCIAL_DATA={payload};\n', encoding='utf-8')
print(f'{len(data)} linhas EA/CW gravadas em {TARGET.name}')
